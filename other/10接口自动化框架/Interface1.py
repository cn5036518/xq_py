#!/usr/bin/env python
#-*- coding:utf-8 -*-

import json
import openpyxl
from Http_request1 import Http_request2

# 读取excel
path1 = r"D:\xq_py\10接口自动化框架\testcase.xlsx"
wb1 = openpyxl.load_workbook(path1)
# print(wb1)
sh1 = wb1.get_sheet_by_name("Sheet1")
max_row1 = sh1.max_row
# print(max_row1)

#2接口类
class Interface2:
    def test(self,url,headers1,params1,method1,data_type1,yes_no1,check_point1,i):
        if method1 == "POST" and data_type1 == "Json":
            h1 = Http_request2(url,params1,headers1)
            result1 = h1.post_request1()
            print(result1)

        ret1 = result1["error_code"]
        ret2 = check_point1["error_code"]

        result1 = json.dumps(result1,ensure_ascii=False)
        if ret1 == ret2:
            print("测试通过")
            sh1.cell(row=i,column=11).value = "通过"
            sh1.cell(row=i, column=12).value = result1
            wb1.save(path1)
        else:
            print("测试不通过")

class For_case:
    def for_case(self):
        for i in range(max_row1-1):
            url1 = sh1.cell(row=i+2,column=3).value  #py3.3只支持这个
            # url1 = sh1.cell(3,3).value   #py3.7可以支持这个
            # print(url1)

            url2 = sh1.cell(row=i+2,column=4).value
            url = url1 +url2

            method1 = sh1.cell(row=i+2,column=5).value

            data_type1 = sh1.cell(row=i+2,column=6).value

            params1 = sh1.cell(row=i+2,column=7).value
            params1 = json.loads(params1)

            check_point1 = sh1.cell(row=i+2,column=8).value
            check_point1 = json.loads(check_point1)

            yes_no1 = sh1.cell(row=i+2,column=10).value
            if yes_no1 =="否":
                continue

            headers1 ={}

            i1 = Interface2()  #新建接口类对象
            i1.test(url,headers1,params1,method1,data_type1,yes_no1,check_point1,i+2) #
            #注意 i+2通过这里传

# 4新建循环case类对象
f1 = For_case()
f1.for_case()

"""
总结：
1、构造方法不是必须的（当普通方法是多个的时候，才需要用到构造方法）
2、读取excel的类中，调接口类，实际参数除了读取的字段，还加一个i+2
3、接口类的构造方法删除，在普通方法test中添加形参
"""















