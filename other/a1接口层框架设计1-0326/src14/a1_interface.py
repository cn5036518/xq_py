#!/usr/bin/env python

"""
整体思路：
1、读取excel的字段
2、发送请求，得到响应数据（调http请求类）-1
3、将响应数据的error_code和exel获取的检查点进行比对-1
    相等，写入通过和响应数据-1
    不相等，写入不通过和响应数据-1
    注意：写入响应数据前，想将python类型转换成json字符串才能写入

方式：1先写死（全局变量来定义），先跑通--1
    2再写活（从excel中读取）--nok
        1、每获取一行，就调一次http请求
        2、先只有一行1，然后循环遍历多行1
    3写入excel
        1、先写入一行1，然后循环遍历写入多行1

一、调用关系
    1、http请求类单独一个文件-1
    2、接口类和读取execel在同一个文件
    3、接口类会调用http请求类（前提：把请求类导入）1
        1、如果请求方法是post并且请求类型是form
            调post方法
            如果请求方法是post并且请求类型是json
            调post方法
            如果请求方法是get
            调get方法
            其他else
    4、读取excel中，会调用接口类

二、接口类
    1、发送请求，得到响应数据（调http请求类，先导入请求类）

三、读取excel
    目的：将读取的excel字段，返回给接口类作为入参
    方式：会调用接口类，传入实际参数
"""

from http_class2 import Http_request1  #导入请求类文件的请求类
import openpyxl
import json

# url="http://v.juhe.cn/laohuangli/d"    #老黄历聚合网站  #这个从excel中读取  先写死
# param1 = {"key": "e711bc6362b3179f5a28de7fd3ee4ace", "date": "2016-5-14"}   #key和value #这个从excel中读取
# headers1 = {}
# method1 = "POST"
# # method1 = "GET"
# data_type1 = "Json"
# # data_type1 = "Form"
# check_point1 = {"error_code":0}


class Interface1:
    def __init__(self,url,param1,headers1,method1,data_type1):
        self.url = url
        self.param1 = param1
        self.headers1 = headers1
        self.method1 = method1
        self.data_type1 = data_type1

    def test(self):
        request1 = Http_request1(self.url, self.param1, self.headers1)
        # 1调http请求类,新建请求类对象 传入成员变量，方便方法间使用
        if self.method1 == "POST" and data_type1 == "Json":
            result1 = request1.post_method1()  #2对象调请求类中的post方法，前提先导入http请求类
            print(result1)
            # print(type(result1))
            # print("----")
        elif self.method1 == "POST" and data_type1 == "Form":
            result1 = request1.post_method1()  # 2对象调请求类中的post方法
            print(result1)
            # print(type(result1))
            # print("----2")
        elif self.method1 == "GET":
            result1 = request1.get_method2()  # 2对象调请求类中的get方法
            print(result1)  #该请求类本身不支持get方法
            print(type(result1))
            print("----3")
        else:
            print("请求失败，请检查用例里的数据是否正确")
        result2 = result1["error_code"]
        # print(result2)  #0
        # print(type(result2))  #int
        check_point2 = check_point1["error_code"]
        #这里的check_point1刚刚读取的时候是json，必须先转换成python字典使用
        # print(check_point2)

        if result2 == check_point2:
            print("错误码是%s，测试通过" % result2)
            #写入单元格
            sh1.cell(row=i+2, column=11).value ="通过"  #注意：将循环遍历中的行号i+2传到这里
            # sh1.cell(row=2, column=12).value = result1
            # 这里的result1是字典，要写入excel，需要先将字典转换成json字符串才行，否则报错
            # result1 = json.dumps(result1,ensure_ascii=False)
            # # ensure_ascii=False  这个参数默认是True，表示只显示ascii，
            # # 如果这个参数改成False，就可以显示non-ascii(包含汉字)
            # sh1.cell(row=i+2, column=12).value = result1
            # wb1.save(path1)
            # 注意：写入单元格后，必须save后，才能生效，没有save的话，写入是没有生效的
            # save的时候，当前的excel文件必须是关闭的，如果是打开的，会出现报错
            # PermissionError: [Errno 13] Permission denied:
            #  'D:\\PycharmProjects\\xiaoqiang\\02base_181027\\001\\lianxi.xlsx'
        else:
            print("错误码是%s，测试不通过" % result2)
            sh1.cell(row=i + 2, column=11).value = "不通过"  #注意：将循环遍历中的行号i+2传到这里
        result1 = json.dumps(result1, ensure_ascii=False)
        sh1.cell(row=i + 2, column=12).value = result1
        wb1.save(path1)

#二 从excel读取字段
#定义路径
path1 = r"D:\PycharmProjects\xiaoqiang\base_181027\a1接口层框架设计1-0326\src14\testcase2.xlsx"

#获取excel文件-工作簿
wb1 = openpyxl.load_workbook(path1)
# print(wb1)

#获取指定sheet-工作表
sh1 = wb1.get_sheet_by_name("Sheet1")
# print(sh1)  #<Worksheet "Sheet1">

#获取最大行号--用于循环遍历
max_row1 =sh1.max_row
# print(max_row1)  #3

#读取单元格的数据
# url1 = sh1.cell(row=2,column=3).value
# cell3 = sh1.cell(row=2,column=2)
# cell2 = sh1.cell(2,2).value
# print(url1)
# print(cell2)
# print(cell3)  #<Cell 'Sheet1'.B2>

for i in range(max_row1-1):   #一共是2行数据，循环2次，循环3次，第三行没有值会报错  必须是max_row1-1
    url1 = sh1.cell(row=i+2, column=3).value
    # print(url1)
    url2 = sh1.cell(row=i+2,column=4).value

    url = url1+url2
    # print(url)

    method1 = sh1.cell(row=i+2,column=5).value
    # print(method1)

    data_type1 = sh1.cell(row=i+2,column=6).value
    # print(data_type1)

    param1 = sh1.cell(row=i+2,column=7).value  #1读取的是json字符串，必须先把json字符串转换成python字典
    # print(param1)
    # print(type(param1)) #<class 'str'>
    param1 = json.loads(param1)  #json字符串转换成python字典
    # print(param1)
    # print(type(param1)) #<class 'dict'>

    check_point1 = sh1.cell(row=i+2,column=8).value  #2读取的是json字符串，必须先把json字符串转换成python字典
    # print(check_point1)
    # print(type(check_point1))  #<class 'str'>
    check_point1 = json.loads(check_point1)  #json字符串转换成python字典
    # print(check_point1)
    # print(type(check_point1))  #<class 'dict'>

    headers1 = {}

    #新建对象
    i1 = Interface1(url,param1,headers1,method1,data_type1)   #注意：这里的param1必须是字典才行，不能是json字符串
    #对象调方法
    i1.test()














