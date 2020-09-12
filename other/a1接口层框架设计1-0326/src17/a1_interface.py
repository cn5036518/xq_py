#!/usr/bin/env python

"""
"""
"""
整体思路：
1、读取excel的字段，将读取的字段作为入参
2、发送http请求给服务器，得到响应数据后
3、将响应数据的错误码和excel中读取的检查点进行比对
4、相等，则往excel写入通过和响应数据
   不相等，则往excel写入不通过和响应数据

整体结构：
1、分为2个文件
    文件1：http请求类1
    文件2：接口类（调用http请求类，
        接口类的形参是读取后的字段，
        判断post或者get以及json或者form，（字段先写死1，然后写活-从excel读取1）
        (先写读1行的时候-1，再写循环遍历多行1)
        将响应数据和检查点进行比对，相等写入通过excel，不相等写入不通过到excel）
            和遍历excel读取字段（通过调用接口类，传入读取后的字段作为实参）
        最后，判断yes_no
注意点：
1、http请求类，返回的是json字符串，将json字符串转换成python字典返回 dic1 = r.json()  1
2、从excel读取的字段是json字符串，在发请求前，必须将json字符串转换成python
    字典才行 param1和check_point1    param1 =json.loads(param1)
    check_point1 = json.loads(check_point1)
3、将响应数据写入excel的时候，调用http请求类，本来返回的是json字符串，
    通过dic1 = r.json()转换成字典后（方便在内存中使用），
    这里需要将python字段转换成json字符串才能写入result1 =json.dumps(result1,ensure=False)
    参数2用于解决写入中文乱码的问题
4、如何保证读取的行和写入的行是同一行？
    遍历循环的i 从读取excel出传递到写入excel的行--关键点
"""

from http_class1 import Http_request1    #从文件导入类
import json
import openpyxl

# url="http://v.juhe.cn/laohuangli/d"
# method1 = "POST"
# # method1 = "GET"
# data_type1 = "Json"
# # data_type1 = "Form"
# check_point1 = {"error_code":0}
# params1 = {"key": "e711bc6362b3179f5a28de7fd3ee4ace", "date": "2016-5-14"}
# yes_no1 = "是"  #扩展1
# headers1 = {}

class Interface1:
    def __init__(self,url,method1,data_type1,params1,check_point1,yes_no1):
        self.url = url   #成员变量，便于方法间使用
        self.method1 = method1
        self.data_type1 = data_type1
        self.params1 = params1
        self.check_point1 =check_point1
        self.yes_no1 = yes_no1

    def test(self):
        h1 = Http_request1(self.url,self.params1,headers1) #1调导入文件，新建对象，这里传入3个参数
        if self.method1 =="POST" and self.data_type1 == "Json":
            result1 = h1.post_request1() #2通过导入文件的新建对象，调方法，这里就不用传参数，新建对象的时候已经传了
            print(result1)
            # print(type(result1))  #<class 'dict'>
        elif self.method1 =="POST" and self.data_type1 == "Form":
            result1 = h1.post_request1()  # 通过导入文件的新建对象，调方法，这里就不用传参数，新建对象的时候已经传了
            print(result1)
        elif self.method1 =="GET":
            result1 = h1.get_request1()  # 通过导入文件的新建对象，调方法，这里就不用传参数，新建对象的时候已经传了
            print(result1)
        else:
            print("请检查测试用例的数据是否准确")

        # print(result1["error_code"])
        error_code1 = result1["error_code"]
        error_code2 = self.check_point1["error_code"]
        #如果self.check_point1是json字符串而不是python字典  会报错
        # TypeError: string indices must be integers

        # if result1["error_code"] == self.check_point1["error_code"]:

        result1 = json.dumps(result1,ensure_ascii=False)  #把字典转换成json字符串
        #注意，ensure_ascii默认是True，即只显示ascii，不显示汉字；改成False就可以显示汉字了
        # if self.yes_no1 == "否":
        #     print("不运行的用例结果置空")
        #     sh1.cell(i + 2, 11).value = ""  # 1写入单元格  #这里的i+2是行数--关键2
        #     # 从而保证下面读取的行，后要写入的行是同一行
        #     sh1.cell(i + 2, 12).value = ""  # 2写入的必须是json字符串，而不能是字典
        #     wb1.save(path1)
        # else:
        if error_code1 == error_code2:
            print("测试通过")
            #写入通过到单元格，写入响应数据到单元格
            sh1.cell(i+2, 11).value = "通过"  #1写入单元格  #这里的i+2是行数--关键2
            #从而保证下面读取的行，后要写入的行是同一行
            sh1.cell(i+2, 12).value = result1  #2写入的必须是json字符串，而不能是字典
            wb1.save(path1) #2写入后，必须保存才能生效；当excel是打开的时候，无法写入
            # 会报错 PermissionError: [Errno 13] Permission denied:
        else:
            print("测试不通过")
            sh1.cell(i+2, 11).value = "不通过"  # 1写入单元格
            sh1.cell(i+2, 12).value = result1  # 2写入的必须是json字符串，而不能是字典
            wb1.save(path1)

#定义excel的路径
path1 = r"D:\PycharmProjects\xiaoqiang\base_181027\a1接口层框架设计1-0326\src17\testcase2.xlsx"

#读取excel的工作簿
wb1 = openpyxl.load_workbook(path1)
# print(wb1)
# print(type(wb1))

#获取excel的工作表Sheet1  #注意Sheet1的S是大写
try:
    sh1 = wb1.get_sheet_by_name("Sheet1")
    # print(sh1)
    # print(type(sh1))
except KeyError as e:
    print(str(e))  #'Worksheet Sheet2 does not exist.'
except Exception as e:
    print(str(e))

#最大行数
max_row1 = sh1.max_row  #3 这里的max_row后面没有小括号，是属性，而不是方法
# print(max_row1) #3

#读取单元格

for i in range(max_row1-1):  #这里的i是行数 ，从第二行开始
# for i in range(max_row1):  #这里的i是行数 ，从第二行开始，第四行是空白，所以 max_row1-1
    url1 = sh1.cell(i+2,3).value  #参数1是行 参数2是列  #这里i+2是行号
    # print(url1)  #http://v.juhe.cn

    url2 = sh1.cell(i+2,4).value  #参数1是行 参数2是列
    # print(url2)  #http://v.juhe.cn

    # url = url1+url2  #http://v.juhe.cn/laohuangli/d
    url = "%s%s" % (url1,url2)  #
    # print(url)

    method1 = sh1.cell(i+2,5).value  #参数1是行 参数2是列
    # print(method1)

    data_type1 = sh1.cell(i+2,6).value  #参数1是行 参数2是列
    # print(data_type1)

    params1 = sh1.cell(i+2,7).value  #参数1是行 参数2是列
    # print(params1)
    # print(type(params1)) #<class 'str'>  从excel读取出来的是json字符串，使用前需要转换成python字典（放入内存）

    params1 =json.loads(params1)  #json字符串转换成字典,否则，会出现--错误的请求KEY
    # print(type(params1))

    check_point1 = sh1.cell(i+2,8).value  #参数1是行 参数2是列
    # print(check_point1)
    # print(type(check_point1)) #<class 'str'> 从excel读取出来的是json字符串，使用前需要转换成python字典（放入内存）

    check_point1 =json.loads(check_point1) #json字符串转换成字典
    # print(type(check_point1))

    yes_no1 = sh1.cell(i+2,10).value
    # print(yes_no1)
    if yes_no1 == "否":
        continue  #跳出当次迭代，进入下一次迭代-小循环

    headers1 = {}

    #新建对象，自动调构造方法 (注意：每获取一行，就调一次接口类)
    i1 = Interface1(url,method1,data_type1,params1,check_point1,yes_no1)

    #对象调方法
    i1.test()















