#!/usr/bin/env python

"""
一、整体思路
    1、读取excel表格的字段
    2、将读取的字段作为入参，发送请求给服务器，获取响应数据
    3、将响应数据的错误码--实际结果
        和读取的检查点的错误码（预期结果）进行比对
    4、相等，写入通过和响应数据到excel表格
       不相等，写入不通过和响应数据到excel表格

二、整体结构
    1、接口类，调http请求类，先导入请求类
        接口类，通过请求方法（POST还是GET）、请求类型（json还是Form）分支来判断，调不同的http请求类的方法
    2、每读取一行，就掉一次接口类（http请求类）--for

三、伪代码
    1、导包
        导入http请求类
        导入json
        导入openpyxl
    2、请求类
        1、调用请求类方法
            通过请求方法（POST还是GET）、请求类型（json还是Form）分支来判断，调不同的http请求类的方法
        2、比对响应数据的错误码和检查点
        3、相等，写入通过和响应数据到excel
           不相等，写入不通过和响应数据到excel
    3、循环for最大行数，获取excel的字段
        每获取一行，就调用接口类一次（http请求类），发一次请求

四、注意点
    1、从excel表格读取后的字段是json字符串，请求数据和检查点都需要转换成python的字典，放入内存使用
        json.loads(params1)
    2、发送http请求给服务器后，获取的响应数据（http请求类返回的是字典），将响应数据写入excel之前，需要
        先将字典转换成json字符串才能写入 json.dumps(响应数据,ensure_ascii=False)
    3、响应数据写入excel后的汉字乱码问题
        通过json.dumps(响应数据,ensure_ascii=False)的参数2解决，默认ensure_ascii=True是不显示汉字的
        改成False后，就可以显示汉字了
    4、写入excel表后，需要保存wb1.save(path1)才能写入生效
        写入的时候，excel表格需要关闭，不能是打开的，否则会包权限错误
    5、如何确保读取的行和写入的行保持一致，最大行数没有小括号
        通过for循环的i进行传递
    6、是否运行，continnue来控制
五、步骤
    1、先将url、params1、headers1、请求方法、请求类型、检查点通过全局变量写死1
    2、然后从excel读取一行，写入一行到excel
    3、最后从excel读取多行-for循环
        写入多行
"""

import json
import openpyxl
from http_class2 import Http_request1

#一、全局变量
# url = "http://v.juhe.cn/laohuangli/d"
# params1 = {"key": "e711bc6362b3179f5a28de7fd3ee4ace", "date": "2016-5-14"}
# headers1 ={}
# method1 = "POST"
# data_type1 = "Json"
# check_point1 = {"error_code":0}
# yes_no1 = "是"

class Interface1:
    def __init__(self,url,params1,headers1,method1,data_type1,check_point1,yes_no1):
        self.url = url  #成员变量，用于方法间使用
        self.params1 = params1
        self.headers1 = headers1
        self.method1 = method1
        self.data_type1 = data_type1
        self.check_point1 = check_point1
        self.yes_no1 = yes_no1

    def test(self):
        #调接口类，新建接口类的对象，传入参数 前提：先导入接口类文件
        h1 = Http_request1(self.url, self.params1, self.headers1)
        if self.method1 == "POST" and self.data_type1 == "Json":
            result1 = h1.post_request1()
            print(result1)
        elif self.method1 == "POST" and self.data_type1 == "Form":
            result1 = h1.post_request1()
            print(result1)
        elif self.method1 == "GET":
            result1 = h1.get_request2()
            print(result1)

        ret1 = result1["error_code"]
        ret2 = self.check_point1["error_code"]
        if ret1 == ret2:
            print("测试通过")
            #写入测试通过和响应数据
            sh1.cell(i+2,11).value = "通过"  #这里是=，而不是==  注意  #这里的i+2就可以保障读取的行和写入的行是同一行
            # #写入响应数据到excel前，先把响应数据（http请求类的return是字典），从字典转换成json字符串才行
            # result1 = json.dumps(result1,ensure_ascii=False)  #参数2是用来解决中文乱码
            # sh1.cell(2,12).value = result1
        else:
            print("测试不通过")
            # 写入测试通过和响应数据
            sh1.cell(i+2, 11).value = "不通过"  # 这里是=，而不是==  注意
            # 写入响应数据到excel前，先把响应数据（http请求类的return是字典），从字典转换成json字符串才行
        result1 = json.dumps(result1, ensure_ascii=False)  # 参数2是用来解决中文乱码
        sh1.cell(i+2, 12).value = result1
        wb1.save(path1)

#二、读取excel的字段
path1 = r"D:\PycharmProjects\xiaoqiang\base_181027\a1接口层框架设计1-0326\src19\testcase2.xlsx"

#1、获取工作簿
wb1 = openpyxl.load_workbook(path1)

#2、获取工作表
sh1 = wb1.get_sheet_by_name("Sheet1")

max_row1 = sh1.max_row

#3、获取单元格
for i in range(max_row1-1):  #0,1
    url1 = sh1.cell(i+2,3).value
    url2 = sh1.cell(i+2,4).value
    url = url1+url2
    # print(url)

    method1 = sh1.cell(i+2,5).value
    data_type1 = sh1.cell(i+2,6).value
    params1 = sh1.cell(i+2,7).value  #从excel表格获取是json字符串，需要装换成python字典，放入内存，用于发送http请求
    #如果是json字符串，而不是字典，发送http请求就会报错
    params1 = json.loads(params1)

    check_point1 = sh1.cell(i+2,8).value  #从excel表格获取是json字符串，需要装换成python字典，放入内存，用于获取错误码
    check_point1 = json.loads(check_point1)

    yes_no1 = sh1.cell(i+2,10).value
    if yes_no1 == "否":
        continue

    headers1 = {}

    #新建接口类对象，自动调构造方法，传入参数
    i1 = Interface1(url,params1,headers1,method1,data_type1,check_point1,yes_no1)
    i1.test()
















