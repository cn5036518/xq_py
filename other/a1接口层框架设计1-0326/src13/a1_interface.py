#!/usr/bin/env python

"""
设计接口自动化测试框架的步骤
1、整体构思-中心思想和提纲（不考虑具体的实现，即作文内容--和写作文的类比）
    比如：流程图 结构图 文字  （听的时候，思考少；写的时候，思考多）
2、目录规划
3、伪代码
4、写代码（硬编码，先能跑起来，小步迭代  比如：参数写死，功能最少demo）

一、整体构思
1主入口--2遍历接口--3发送请求--4得到返回--5对比结果--6测试报告--7发送邮件

二、目录结构
1、主入口的py文件--main
2、核心代码py文件夹--src
    遍历接口
    发送请求-得到返回
    对比结果

3、存放接口测试用例的文件夹--case
   存放形式：excel（推荐）或者数据库
   推荐excel的原因：
   1、维护excel简单
   2、数据库学习成本稍高
   3、简便性上excel占优

   excel的字段设计：
   1、url
   2、请求报文
   3、请求方式post还是get
   4、实际返回报文
   5、期望结果
   6、通过状态
   7、备注

4、存在公共函数的文件夹--commons
   公共函数：已经封装好的，比较独立的实现一个小功能，且不依赖别的部分
   例如：http的请求类
        发送邮件的类
        发日志的类

5、存放测试报告的文件夹--reports
   excel形式的报告（推荐）优于html的报告
   原因：
   1、excel的统计功能

6、放插件的文件夹--others
7、存放日志文件的文件夹--logs

三、课上练习
1、目前先把所有的py文件全不放在src核心文件夹
    等代码调通后，最后将py文件分发到其他目录中，从而避免开始就出现导包错误
2、http请求类就绪
3、接口请求类的编写：class Interface_test:
    调用请求类并对结果进行对比判断
    不用unittest的断言，而是使用re.search完成
    所有参数全部写死（不从excel进行读取）
4、保证上面的代码可以正常运行并且测试通过

四、课上练习伪代码
    class Interface_test:
        def __init__(self,参数列表):
            self.url = url
            self.params1 = params1
            self.headers1 = headers

        def 方法名(self):
            http请求类的调用
            发送请求
            获取响应

            if self.method1 = "post":
                if 数据类型是Form:
                    re = http.post_form(参数列表)
                elif 数据类型是Json:
                    header = {"Content-Type":"application/json;charset=utf8"}
                    re =http.post_json(参数列表)
            elif self.method1 ="get":
                接收响应数据
                re =http.get1(参数列表)
            else:#主要是异常处理
                # 记录日志文件
                print("请求失败，请检查用例里的数据是否正确")

            re.search(检查点,响应内容)
            if re.search(error_code="0",请求类的整个返回r):
            if re.search("0",str(字典(["error_code"])))):  #参数1是检查点
            #如果这个正则忘了-不会，不写出来，后面的代码无法运行，
                就可以把正则直接写成1-True，先让代码继续，回头再来完善这个正则，就不会卡在这里了
                print("请求成功")
            else:
                print("请求失败")

#search函数 匹配字符串的全部，并且返回第一个符合条件的字符串（从左到右）
print(re.search("www","22wwwcom").group()) #www  扫描字符串全部
# 参数1是正则表达式，参数2是待匹配的字符串  group()是显示匹配后的结果,可视化

五、
1、从读取一条记录，--ok
   然后遍历多行，并且是每遍历一次，就调一次请求
2、先完成post请求，再完成判断，如果post请求，就如何，如果get 请求就如何--ok

六、遍历excel
1、打开excel
2、获取指定的sheet
3、for循环遍历case
   01是否运行处，如果是yes就运行，如果是no就不允许--ok
   （if 是否运行处取出的单元格!=yes continue,跳出当次迭代，继续下一次迭代）
   02除了测试结果和响应数据之外--最后两列之外，其余的字段都需要取出来
      因为最后两列的需要往里面填写数据--写excel，所以不需要读取

七、注意
把testcase的检查点0改成{"error_code":0}--
注意：写法只能是{"error_code":0}，0是数字，且是json格式（有大括号，没有大括号就不行）

"""

import re
from http_class2 import Requests_Post1   #导入http请求文件的类-post类
from http_class2 import Requests_Get1   #导入http请求文件的类--get类
import openpyxl
import json

#2 定义接口类
class Interface_test:
    def __init__(self,url,param1,headers,check_point1,method1,data_type1,row):
        self.url = url
        self.params1 = params1
        self.headers1 = headers  #传入参数，成员变量用于方法间调用
        self.check_point1 = check_point1  #检查点（期望结果 比如 error_code = "0"# ）
        self.method1 = method1  #post还是get方法
        self.data_type = data_type1
        self.row = row  #行号  注意点：保证读取行和写入行是同一行

    def test(self):
        # 2新建对象
        if self.method1 == "POST" and self.data_type == "Form":
            requests1 = Requests_Post1(self.url, self.params1,self.headers1)  #
            #调http请求类
            # 传入实际参数
            # 3通过对象调方法
            result1 = requests1.requests_post2()
            print("post请求-form获取响应结果（python-字典类型）=", result1)  # 5输出整个返回内容（字典类型）
        elif self.method1 == "POST" and self.data_type == "Json":
            requests1 = Requests_Post1(self.url, self.params1, self.headers1)  # 传入实际参数
            # 3通过对象调方法
            result1 = requests1.requests_post2()
            print("post请求-json获取响应结果（python-字典类型）=", result1)  # 5输出整个返回内容（字典类型）
        elif self.method1 == "GET":
            requests1 = Requests_Get1(self.url, self.params1, self.headers1)  # 传入实际参数
            # 3通过对象调方法
            result1 = requests1.requests_get2()
            print("get请求获取响应结果（python-字典类型）=", result1)  # 5输出整个返回内容（字典类型）
        else:
            #记录日志文件，异常处理
            print("请求失败，请检查用例里的数据是否正确")

    #二，前面先做分支判断，得到响应数据
        print("错误码error_code是：", result1["error_code"])  # 6输出错误码error_code
        result2 = result1["error_code"]

        if result2 == self.check_point1["error_code"] :   #
            print("post请求返回的error_code是%s，测试通过" % result2)
            # sh1.cell(row=3, column=11).value = "通过"    #行号写死
            sh1.cell(row=self.row, column=11).value = "通过"
            #注意：这里的result1是字典，往excel中写入前，必须将字典转换成json字符串才行，否则报错
            # ValueError: Cannot convert 。。。  to Excel
            # json.loads(json_str)            json字符串转换成字典
            # json.dumps(dict)            字典转换成json字符串
            result1 = json.dumps(result1,ensure_ascii=False)  #将字段转换成json字符串
            #注意2：ensure_ascii默认是True，就不显示汉字，值显示ascii码；
            # 如果ensure_ascii=False，就不仅仅是显示ascii码了，可以显示汉字了
            sh1.cell(row=self.row, column=12).value = result1
            wb1.save(path1)
        else:
            sh1.cell(row=self.row, column=11).value = "不通过"
            result1 = json.dumps(result1, ensure_ascii=False)  # 将字段转换成json字符串
            sh1.cell(row=self.row, column=12).value = result1
            wb1.save(path1)
            print("返回的error_code是%s，测试不通过" % result2)

# 全局变量
path1 = r"D:\PycharmProjects\xiaoqiang\base_181027\a1接口层框架设计1-0326\src13\testcase2.xlsx"

# 1 打开工作簿文件
wb1 = openpyxl.load_workbook(path1)  # 1打开excel文件（类似打开文件对象）
# print("获取excel文件-工作簿的所有工作表名字",wb1.get_sheet_names()) #2返回工作表的名字作为列表的元素

# 2 获取指定工作表
sh1 = wb1.get_sheet_by_name("Sheet1")  # 3指定要操作的工作表sheet1
# print("获取指定工作表的名字",sh1.title) #3获取指定工作表的名字 Sheet1

# #003 获取最大的列数和最大的行数
max_row1 = sh1.max_row
# print("获取最大行数",max_row1)    #3
# max_column1 = sh1.max_column
# # print("获取最大列数",sh1.max_column)  #12

for i in range(max_row1-1):  # 获取每一行的数据
    # 注意从单元格取出来的数据都要replace('\n','').replace('\r','')   去掉空格或者换行
    url1 = sh1.cell(row=i + 2, column=3).value
    # print(url1)
    url2 = sh1.cell(row=i + 2, column=4).value
    # print(url2)
    url= url1+url2
    # print(url)
    method1 = sh1.cell(row=i + 2, column=5).value  #从第二行开始
    # print(method1)
    data_type1 = sh1.cell(row=i + 2, column=6).value
    # print(data_type1)
    params1 = sh1.cell(row=i + 2, column=7).value#特别注意:从excel读取的请求数据是
    # '{"key": "e711bc6362b3179f5a28de7fd3ee4ace", "date": "2016-5-14"}',这个是json字符串，而不是python字典
    #  json字符串在python中使用前，必须先将json转换成字典才能使用（否则报错）
    #  json字符串  '{"key": "e711bc6362b3179f5a28de7fd3ee4ace", "date": "2016-5-14"}'
    #  字典         {"key": "e711bc6362b3179f5a28de7fd3ee4ace", "date": "2016-5-14"}

    # JSON到字典转化：
    params1 = json.loads(params1)
    # print(params1)
    # print(type(params1))

    check_point1 = sh1.cell(row=i + 2, column=8).value
    # print(check_point1)
    # print(type(check_point1))  #<class 'str'>
    # 特别注意:从excel读取的请求数据是 '{"error_code":0}'，这个是json字符串，而不是python字典
    # json字符串在python中使用前，必须先将json转换成字典才能使用（否则报错）
    # JSON到字典转化：
    check_point1 = json.loads(check_point1)
    # print(type(check_point1)) #<class 'dict'>

    yes_no1 = sh1.cell(row=i + 2, column=10).value
    # print(yes_no1)
    # print(type(yes_no1))

    if yes_no1 != "是":
        continue #跳出当次迭代，继续下一次迭代

    headers1 = {}

    test1 = Interface_test(url, params1, headers1, check_point1, method1, data_type1,i+2)
    #2 调接口类
    # 自动执行构造方法，传入参数
    #注意：这里的行号必须是i+2 而不是i或者i+1  i+2的话，第一次循环是2，第二次循环是3，for只循环2次
    #     方法论的关键是 将i+2的值打印出来，可见不跳步骤，一步一步打印的重要性+耐心
    # # 对象调方法
    test1.test()

"""
写代码方法：
1、整理思路--写作文的中心思想
2、写伪代码--列作文提纲
3、将伪代码翻译成镇代码--将提纲丰富成作文内容

一、整理思路
    1、读取excel中的字段
    2、发送请求，得到响应数据
    3、将响应数据和检查点-期望结果error_code进行比对
    4、相等，则写入通过到excel，并写入响应数据
       不相等，则写入不通过到excel，并写入响应数据

二、整理结构
    1、http请求类单独出来，作为一个文件
    2、读取excel中的字段
        1、目的：将读取的每一行作为请求类的入参
        2、方式：循环遍历整个excel的行，每读取一行，就调用一次请求类
        3、注意点：读取的单元格是json字符串，需要转换成python字典，来使用
    3、每发送一次请求，就得到一个响应数据
    4、将响应数据的error_code和检查点进行比对
    5、相等，则写入通过到excel，并写入响应数据
       不相等，则写入不通过到excel，并写入响应数据
        1、注意点：写入通过或者响应数据前，需要把带写入的内容转换成json字符串才行，否则报错

三、伪代码
 1、http请求类单独成一个文件
    1、该类包含2个方法，post请求的处理和get请求的处理
    2、通过requests.post或者requests.get来发送请求
    3、注意点：
        1、捕捉异常
        2、返回的响应数据是json字符串，在这里就转换成python字典，从而放在内存中使用
            dic1 = r.json()
 2、读取excel中的字段
    1、目的：将读取的每一行作为请求类的入参
    2、方式：循环遍历整个excel的行，每读取一行，就调用一次请求类
    3、注意点：读取的单元格是json字符串，需要转换成python字典，来使用
    4、实现思路：
        全局变量
        1、指定excel路径
        2、打开excel的工作簿文件
        3、获取excel的制定工作表，比如“sheet1”

        1、获取excel的行数max_row
        2、for循环遍历整个行数
        3、获取url、请求数据、请求方法post还是get、请求类型json还是form
            检查点error_code、是否运行
            上述字段的列是固定的，行号是i
        4、每读取一行，就调用一次接口类

        二、接口类单独封装出来，在同一个文件
            1、判断如果请求方法是post且请求类型是json
                就调用请求类，入参是读取的字段，通过构造方法，成员变量来传递
                得到响应数据
            2、判断如果请求方法是post且请求类型是form
                就调用请求类，入参是读取的字段，通过构造方法，成员变量来传递
                得到响应数据
            3、判断如果请求方法是get
                就调用请求类，入参是读取的字段，通过构造方法，成员变量来传递
                得到响应数据
            4、将上述响应数据从json字符串转换成python字典
            5、获取字典中的error_code
            6、对比响应数据中的error_code和excel中读取的检查点的error_code是否相等
            7、相等，就写入通过和响应数据
               不相等，就写入不通过和响应数据
               注意点:写入前，把字典转换成json字符串后，才能写入

三、调用关系
    1、http请求类单独一个文件
    2、接口类和读取execel在同一个文件
    3、接口类会调用http请求类（前提：把请求类导入）
    4、读取excel中，会调用接口类
"""






















