#!/usr/bin/env python


"""
一、思路
    1、读取excel的一行字段
    2、将url和请求数据作为入参，调用http请求类，给服务端发送请求，获取响应数据
    3、将响应数据的错误码--实际结果
       和检查点的错误码期望结果--进行比对
    4、相等，则写入通过和响应数据
        不相等，则写入不通过和响应数据

二、结构
    1、接口类
        调http请求类
        进行比对判断
        相等，写入通过和响应数据
        不相等，写入不通过和响应数据
    2、循环遍历excel，读取excel的字段
        调接口类，传入实际参数（excel的字段）

三、伪代码
    1、导包
        json
        openpyxl
        http请求类
    2、接口类
        1、判断请求方法是post还是get，请求类型是json还是form
            分别调http请求类的不同方法，获取响应数据（字典）
        2、将响应数据的错误码和检查点的错误码比对
           1、相等，写入通过和响应数据
           2、不相等，写入不通过和响应数据
    3、循环读取excel
    4、for中调用接口类，传入实际参数

四、注意点
    1、读取excel后，获取的请求数据和检查点是json字符串，需要转换成python字典，放入内存
        json.loads(params1)
    2、写入excel响应数据的时候，响应数据是从http请求类获取的字典，需要先把字典转换成json字符串，才能写入
        json.dumps(响应数据,ensure_ascii = False) 参数2是用来解决中文乱码问题，默认ensure_ascii = True是不显示中文的
    3、写入后，需要保存才能生效wb1.save(path1)
    4、保存的时候，excel不能是打开的，否则，会出现权限不足
    5、如何保证读取的行和写入的行是同一行，通过i+2来实现

五、步骤
    1、先全局变量写死1
    2、然后通过excel只读取一行，写入一行
    3、最后通过excel读取多行，写入多行

六、扩展和优化方法
    1、错误key
    2、出现500 404 403 如何捕捉异常，将状态码写入表格
"""

#一、导包
import json
import openpyxl
from http_class2 import Http_request  #导入http请求类

# #二、全局变量
# url = "http://v.juhe.cn/laohuangli/d"
# params1 = {"key": "e711bc6362b3179f5a28de7fd3ee4ace", "date": "2016-5-14"}
# headers1 = {}
# method1 = "POST"
# data_type1 = "Json"
# check_point1 = {"error_code":0}
# yes_no1 = "是"

#三、接口类
class Interface1:
    def __init__(self,url,params1,headers1,method1,data_type1,check_point1,yes_no1):
        self.url = url  #成员变量，方法间使用
        self.params1 = params1
        self.headers1 =headers1
        self.method1 = method1
        self.data_type1 = data_type1
        self.check_point1 = check_point1
        self.yes_no1 = yes_no1

    def test(self):
        #新建http请求类对象，前提是导入请求类
        h1 = Http_request(self.url, self.params1, self.headers1)
        if self.method1 == "POST" and self.data_type1 == "Json":
            result1 = h1.post_request1()  #必须有小括号 注意
            print(result1)
        elif self.method1 == "POST" and self.data_type1 == "Form":
            result1 = h1.post_request1()  # 必须有小括号 注意
            print(result1)
        elif self.method1 == "GET":
            result1 = h1.get_request2()  # 必须有小括号 注意
        else:
            print("请检查测试用例数据")

        ret1 = result1["error_code"]
        ret2 = self.check_point1["error_code"]
        result1 = json.dumps(result1, ensure_ascii=False)  # 参数2是解决写入中文乱码问题 字典转换成json字符串
        if ret1 == ret2:
            print("测试通过")
            #写入通过和响应数据
            sh1.cell(i+2,11).value = "通过"
        else:
            print("测试不通过")
            sh1.cell(i+2, 11).value = "不通过"
            #写入响应数据
        sh1.cell(i+2, 12).value = result1  # 这个是字典，写入excel前，需要将字典转换成json字符串才行
        wb1.save(path1)


#四、读取excel字段
#1 文件路径
path1 = r"D:\PycharmProjects\xiaoqiang\base_181027\a1接口层框架设计1-0326\src20\testcase2.xlsx"

#2 获取工作簿
wb1 = openpyxl.load_workbook(path1)

#3 获取工作表
sh1 = wb1.get_sheet_by_name("Sheet1")

#3行数
max_row1 = sh1.max_row  #没有小括号，属性

#4 获取单元格

for i in range(max_row1-1):  #0,1
    url1 = sh1.cell(i+2,3).value
    # print(url1)

    url2 = sh1.cell(i+2,4).value
    url = url1 + url2

    method1 = sh1.cell(i+2,5).value
    data_type1 = sh1.cell(i+2,6).value

    params1 = sh1.cell(i+2,7).value   #从exccel获取的是json字符串，发送请求的时候，需要转换成python字典，
    params1 =json.loads(params1)

    check_point1 = sh1.cell(i+2,8).value  #从exccel获取的是json字符串，发送请求的时候，需要转换成python字典，
    check_point1 =json.loads(check_point1)

    yes_no1 = sh1.cell(i+2,10).value
    if yes_no1 == "否":  #跳出当次迭代，进入下一次迭代（小循环）
        continue

    # print(yes_no1)
    headers1 = {}   #注意 不要遗漏

    #五、新建接口类对象，传入参数，自动调构造方法 注意：调接口类在for循环里面，从而实现读取一行，就掉一次请求
    i1 = Interface1(url,params1,headers1,method1,data_type1,check_point1,yes_no1)

    #对象调方法
    i1.test()

























