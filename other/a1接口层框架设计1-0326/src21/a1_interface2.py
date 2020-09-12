#!/usr/bin/env python


"""
一、思路
    1、从excel读取字典，作为请求类的参数url params1 headers1，发送给服务端，获取响应数据
    2、将响应数据的错误码和检查点的错误码比对
    3、相等，将通过和响应数据写入
        不相等，将不通过和响应数据写入excel

二、结构
    1、接口类
    2、接口类中要调用Http请求类
    3、要通过判断请求方法是post还是get，请求类型是json还是form，调用请求类，等到响应数据
    4、将响应数据的错误码和检查点错误码进行比对
    5、相等，写入通过和响应数据
       不相等，写入不通过和响应数据

    6、循环遍历读取excel，获取每个字段
    7、新建接口类对象，调接口类的方法，传入实际参数

三、伪代码
    1、导包
        json
        openpyxl
        Http请求类
    2、全局变量先写死
        url，请求数据，headers1，请求方法，请求类型，检查点，是否运行
    3、接口类
        1、构造函数 传入参数，成员变量
        2、普通方法test
            1、通过判断请求方法是post还是get，请求类型是json还是form，调用请求类，得到响应数据
        3、将响应数据的错误码和检查点错误码进行比对
        4、相等，写入通过和响应数据
           不相等，写入不通过和响应数据
    4、循环遍历读取excel，获取每个字段
        1、指定excel文件路径
        2、获取工作簿
        3、获取工作表
        4、获取单元格
    5、新建接口类对象，调接口类的方法，传入实际参数

四、注意点
    1、读取excel字段的时候，对于请求数据和检查点，读取的是json字符串，需要将json字符串转换成python的字典，放在内存中
        json.loads(params1)
        json.loads(check_point1)
    2、写入excel响应数据的时候，调用请求类，返回的是响应数据是字典，需要将字典转换成json字符串，才能写入
        json.dumps(响应数据,Ensure_ascii = False) 参数2用于显示中文
    3、写入excel后，必须wb1.save(path1),否则写入没有生效
    4、写入excel的时候，表格必须是关闭的，否则会报权限错误
    5、如何保证读取行和写入行是同一行，通过i+2

五、步骤
    1、先全局变量写死，将入参字段1
    2、然后再从excel读取一行，写入一行
    3、最后从excel读取多行，写入多行

六、扩展和优化
    错误key--1
    状态码是500 404 403等
"""

#一 导包
import json
import openpyxl
from http_class2 import Http_request1

# #二 全局变量写死
# url = "http://v.juhe.cn/laohuangli/d"
# params1 = {"key": "e711bc6362b3179f5a28de7fd3ee4ace", "date": "2016-5-14"}
# headers1 = {}
# method1 = "POST"
# data_type1 = "Json"
# check_point1 = {"error_code":0}
# yes_no1 = "是"

#三、接口类
class Interface1:
    def __init__(self,url,params1,headers1,method1,data_type1,check_point1,yes_no1):
        self.url = url
        self.params1 = params1
        self.headers1 = headers1
        self.method1 = method1
        self.data_type1 = data_type1
        self.check_point1 = check_point1
        self.yes_no1 = yes_no1

    def test(self):
        h1 = Http_request1(self.url, self.params1, self.headers1)
        if self.method1 == "POST" and self.data_type1 == "Json":
            result1 = h1.post_request1()   #小括号要加上
            print(result1)
        elif self.method1 == "POST" and self.data_type1 == "Form":
            result1 = h1.post_request1()
            print(result1)
        elif self.method1 == "GET":
            result1 = h1.get_request2()
            print(result1)
        else:
            print("请检查测试用例的数据")

        ret1 = result1["error_code"]
        ret2 = self.check_point1["error_code"]

        if ret1 == ret2:
            print("测试通过")
            #写入通过和响应数据（需要将响应数据从字典转换衬json字符串才能写入）
            sh1.cell(i+2,11).value = "通过"  #通过i+2 保证读取的行和写入的行是同一行
        else:
            print("测试不通过")
            sh1.cell(i+2, 11).value = "不通过"
        result1 = json.dumps(result1,ensure_ascii=False)  #字典转换成json字符串，参数2是显示中文的
        sh1.cell(i+2, 12).value = result1
        wb1.save(path1)


#四、读取excel
#1指定excel的路径
path1 = r"D:\PycharmProjects\xiaoqiang\base_181027\a1接口层框架设计1-0326\src21\testcase2.xlsx"

#2 获取工作簿
wb1 = openpyxl.load_workbook(path1)

#3 获取工作表
sh1 = wb1.get_sheet_by_name("Sheet1")

#3 最大行数
max_row1 = sh1.max_row   #没有小括号，属性

#4 获取单元格
for i in range(max_row1-1):  #0,1

    url1 = sh1.cell(i+2,3).value
    # print(url1)

    url2 = sh1.cell(i+2,4).value
    url = url1 + url2

    method1 = sh1.cell(i+2,5).value

    data_type1 = sh1.cell(i+2,6).value

    params1 = sh1.cell(i+2,7).value
    params1 = json.loads(params1)  #从excel获取的是json字符串，转成hpython字典，放入内存

    check_point1 = sh1.cell(i+2,8).value
    check_point1 = json.loads(check_point1)  # 从excel获取的是json字符串，转成python字典，放入内存

    yes_no1 = sh1.cell(i+2,10).value
    # print(yes_no1)
    if yes_no1 == "否":
        continue  #跳出当次小循环，进行下一次小循环

    headers1 = {}  #容易忘记 注意

    #五、新建接口类对象，传入参数，自动调构造方法  （调接口类方法，必须在for内，每读取一行，发一次请求）
    i1 = Interface1(url,params1,headers1,method1,data_type1,check_point1,yes_no1)
    i1.test()


























