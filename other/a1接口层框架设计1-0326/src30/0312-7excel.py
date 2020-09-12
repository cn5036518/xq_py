#!/usr/bin/env python
# -*- coding:utf-8 -*-

'''
@author: jack
@license: (C) Copyright 2013-2017, Node Supply Chain Manager Corporation Limited.
@contact: cn5036520@163.com
@software: sign
@file: 0312-7excel.py
@time: 2018/12/9 6:54
@desc:
'''

import openpyxl #只支持.xlsx格式的，即支持excel2007,2010及以上版本的，不支持excel2003

#一、openpyxl安装步骤--python下第三方安装包的安装方法（通用的）
"""
1、网上下载openpyxl安装包解压后，存放在本地 F:\百度网盘\03py_xq\day1_170223\openpyxl-2.5.12
2、cmd下进入路径 F:\百度网盘\03py_xq\day1_170223\openpyxl-2.5.12
3、 执行python setup.py install（前提，python3.3的环境变量需要配置好）
4、提示Finished processing dependencies for openpyxl==2.5.12 表示安装成功
5、代码中执行import openpyxl没有报错，表示包已经安装好了（验证方法）
    或者在cmd下中python后，再执行import openpyxl，如果没有报错就表示包安装成功

C:\WINDOWS\system32>python
Python 3.3.5 (v3.3.5:62cf4e77f785, Mar  9 2014, 10:35:05) [MSC v.1600 64 bit (AMD64)] on win32
Type "help", "copyright", "credits" or "license" for more information.
>>> import openpyxl
>>>

python3.3环境变量设置
在环境变量中添加Python目录
(1) 右键点击"计算机"，然后点击"属性"
(2) 然后点击"高级系统设置"
(3) 选择"系统变量"窗口下面的"Path",双击即可！
(4) 然后在"Path"行，添加python安装路径即可(C:\Python33)。
    ps：记住，路径直接用分号"；"隔开！
"""

#二、操作excel
import openpyxl
import time

path1 = r"D:\PycharmProjects\xiaoqiang\02base_181027\001\lianxi.xlsx"
path2 = r"D:\PycharmProjects\xiaoqiang\02base_181027\001\lianxi2.xlsx"

#1 打开工作簿文件
wb1 = openpyxl.load_workbook(path1)  #1打开excel文件（类似打开文件对象）
print("获取excel文件-工作簿的所有工作表名字",wb1.get_sheet_names()) #2返回工作表的名字作为列表的元素
# 2 ['Sheet1', 'Sheet2'] 返回列表

#2 获取指定或者活动工作表
sh1 = wb1.get_sheet_by_name("Sheet1")  #3指定要操作的工作表sheet
print("获取指定工作表的名字",sh1.title) #3获取指定工作表的名字 Sheet2
#
# sh2 = wb1.get_active_sheet()
# print("获取当前活动的工作表名字",sh2)
# 4获取当前活动的工作表名字 <Worksheet "Sheet1"> 即鼠标光标所在地方


#3 操作单元格
#001 读取单元格数据
print("获取指定工作表的A1单元格的值",sh1["A1"].value)  #方法1
print("获取指定工作表的A1单元格的值",sh1.cell(row=1,column=1).value) #方法2 方便for循环取值--推荐
print("获取指定工作表的A1单元格的值",sh1.cell(1,1).value) #方法2 cell的参数1是行，参数2是列
# 1获取指定工作表的A1单元格的值 加

#002 写入单元格数据并保存save
sh1.cell(row=1,column=3).value = "tets33"   #方式1--推荐
# sh1["A4"].value = "test22334"  #方式2
# wb1.save(path1) #参数是当前文件，修改当前文件后，保存（当前文件修改了）
# 注意：写入单元格后，必须save后，才能生效，没有save的话，写入是没有生效的
wb1.save(path2) #参数不是当前文件，修改当前文件后，另存为 （当前文件没有修改）--类似关闭excel对象前，先保存
# 注意：写入单元格后，必须save后，才能生效，没有save的话，写入是没有生效的
#save的时候，当前的excel文件必须是关闭的，如果是打开的，会出现报错
# PermissionError: [Errno 13] Permission denied:
#  'D:\\PycharmProjects\\xiaoqiang\\02base_181027\\001\\lianxi.xlsx'

#003 获取最大的列数和最大的行数
print("获取最大列数",sh1.max_column)  #4
print("获取最大行数",sh1.max_row)    #14

















