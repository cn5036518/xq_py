#!/usr/bin/env python
#-*- coding:utf-8 -*-

import re
import openpyxl
import json

#2 定义遍历case类
class Forcase:
    def forcase(self):
    #二、把参数写活，从excel读取
    # path1 = r"D:\PycharmProjects\xiaoqiang\base_181027\a1接口层框架设计1-0326\testcase.xlsx"
    # path1 = r"D:\PycharmProjects\xiaoqiang\base_181027\a1接口层框架设计1-0326\testcase2.xlsx"
        path1 = r"D:\PycharmProjects\xiaoqiang\base_181027\a1接口层框架设计1-0326\src3\testcase2.xlsx"

        #1 打开工作簿文件
        wb1 = openpyxl.load_workbook(path1)  #1打开excel文件（类似打开文件对象）
        # print("获取excel文件-工作簿的所有工作表名字",wb1.get_sheet_names()) #2返回工作表的名字作为列表的元素

        #2 获取指定工作表
        sh1 = wb1.get_sheet_by_name("Sheet1")  #3指定要操作的工作表sheet1
        # print("获取指定工作表的名字",sh1.title) #3获取指定工作表的名字 Sheet1

        # #003 获取最大的列数和最大的行数

        max_row1 = sh1.max_row
        # print("获取最大行数",max_row1)    #3
        max_column1 = sh1.max_column
        # print("获取最大列数",sh1.max_column)  #12

        #2

        li1 = []
        for i in range(max_row1-1):  # 获取每一行的数据
            # 注意从单元格取出来的数据都要replace('\n','').replace('\r','')   去掉空格或者换行
            url1 = sh1.cell(row=i + 2, column=3).value
            # print(url1)
            url2 = sh1.cell(row=i + 2, column=4).value
            # print(url2)
            url= url1+url2
            # print(url)
            method1 = sh1.cell(row=i + 2, column=5).value  #从第二行开始
            # print(method1)
            data_type1 = sh1.cell(row=i + 2, column=6).value
            # print(data_type1)
            params1 = sh1.cell(row=i + 2, column=7).value#特别注意:从excel读取的请求数据是
            # '{"key": "e711bc6362b3179f5a28de7fd3ee4ace", "date": "2016-5-14"}',这个是json字符串，而不是python字典
            #  json字符串在python中使用前，必须先将json转换成字典才能使用（否则报错）
            #  json字符串  '{"key": "e711bc6362b3179f5a28de7fd3ee4ace", "date": "2016-5-14"}'
            #  字典         {"key": "e711bc6362b3179f5a28de7fd3ee4ace", "date": "2016-5-14"}

            # JSON到字典转化：
            params1 = json.loads(params1)
            # print(params1)
            # print(type(params1))

            check_point1 = sh1.cell(row=i + 2, column=8).value
            # print(check_point1)
            # print(type(check_point1))  #<class 'str'>
            # 特别注意:从excel读取的请求数据是 '{"error_code":0}'，这个是json字符串，而不是python字典
            # json字符串在python中使用前，必须先将json转换成字典才能使用（否则报错）
            # JSON到字典转化：
            check_point1 = json.loads(check_point1)
            # print(type(check_point1)) #<class 'dict'>

            yes_no1 = sh1.cell(row=i + 2, column=10).value
            # print(yes_no1)
            # print(type(yes_no1))

            if yes_no1 != "是":
                continue #跳出当次迭代，继续下一次迭代

            headers1 = {}

            li1.append((url, params1, headers1, check_point1, method1, data_type1, yes_no1))
            # return (url, params1, headers1, check_point1, method1, data_type1,yes_no1)
        return (li1)
        # print(li1)
            # li1 = []
            # li1.append((url, params1, headers1, check_point1, method1, data_type1,yes_no1))


#二、新建遍历类对象
forcase1 = Forcase()
#遍历类对象调方法--目的是拿到excel中的取值，作为元组返回
ret1 =forcase1.forcase()
print(ret1)
# #('http://v.juhe.cn/laohuangli/d', {'key': 'e711bc6362b3179f5a28de7fd3ee4ace', 'date': '2016-5-14'},
# # {}, {'error_code': 0}, 'POST', 'Form', '是')























