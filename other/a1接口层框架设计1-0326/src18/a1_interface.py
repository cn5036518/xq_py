#!/usr/bin/env python

"""
一、整体思路
1、http请求类以及完成，导入请求类
2、读取excel的字段
3、将excel字段作为入参，调用http请求，获取响应数据
4、将响应数据和检查点进行比对
5、相等，写入通过和响应数据到excel
   不相等，写入不通过和响应数据到excel

二、整体结构
1、接口类
    1、通过post get或者Json Form进行判断，如何调http请求类
    2、将响应数据和检查点进行比对

三、伪代码
    1、导包
        导入接口类
        导入json
        导入open--excel包
    2、定义全局变量  url params headers 请求方法 请求类型 检查点 是否执行等
    3、接口类
        构造方法
        普通方法test
        if 请求方法=="POST" and 请求类型 =="Json":
            result1 = 调http请求类的post方法获取响应数据 --字典类型
        elif 请求方法=="POST" and 请求类型 =="Form":
            调http请求类的post方法
        elif 请求方法=="GET" :
            调http请求类的get方法
        else:
            print("请检查测试用例数据")
        if result2["error_code"] == chech_point["error_code"]
            操作excel，将通过写入指定单元格  sh1.cell(行号，列号).value="通过"
            将响应数据转换成json字符串后，写入指定单元格
    4、获取单行的excel的字段
        1、指定excel的路径path1
        2、获取指定的工作簿文件 load workbook  wb1
        3、获取指定的工作表 get sheet by name  sh1
        4、获取指定的单元格（url 请求数据 请求方法 请求类型 检查点 是否运行-continue等）
            sh1.cell(行号，列号).value
    5、读取多行的excel的字段
        1、获取最大行数 max_row
        2、for i in range(max_row)  注意行号
                url = sh1.cell(i，列号).value
                params1 = sh1.cell(i，列号).value
                method1 = sh1.cell(i，列号).value
                data_type1 = sh1.cell(i，列号).value
                check_point1 = sh1.cell(i，列号).value
                yes_no1 = sh1.cell(i，列号).value

    6、调用接口类（新建接口类对象，对象来调方法）
        每读取一行，就调用一次请求

四、注意点
1、从excel读取的字段，是json字符串，params和检查点需要转换成python的字典后使用  json.loads()
2、发送请求后，获取的响应数据是字典（http请求类的return）
    但是往excel中写入响应数据的时候，需要把字段转换成json字符串，才行  json.dumps()
3、写入excel的响应数据-json字符串的汉字乱码问
   json.dumps(Non-ascii=False)  这个non-ascii默认是True，即不显示非ascii，
   改成False即可显示汉字
4、如何保证读取行的行号就是写入行的行号
   通过for循环的i进行传递
5、写入excel的时候
    1、注意save后才能生效
    2、写入保存的时候，excel不能是打开的，否则有权限错误
6、新建对象的时候，传入参数（自动调构造方法）
    对象调方法的时候，就不用再次传入参数了

五、方法总结
1、先用写死的url params headers，调http请求类，获取响应数据1
2、再从excel读取一行，获取响应数据，写入通过和响应数据到excel 1
3、最后，遍历循环读取多行，获取响应数据，写入通过和响应数据到excel 1

"""

import json
import openpyxl
from http_class2 import Http_request  #导入请求类

#一、定义全局变量，写死
# url params headers 请求方法 请求类型 检查点 是否执行等

# url="http://v.juhe.cn/laohuangli/d"
# params1 = {"key": "e711bc6362b3179f5a28de7fd3ee4ace", "date": "2016-5-14"}
# headers1 = {}
# method1 = "POST"
# data_type1 = "Form"
# check_point1 = {"error_code":0}
# yes_no1 = "是"


#二、接口类（调请求类的方法）
class Interface1:
    def __init__(self,url,params1,headers1,method1,data_type1,check_point1,yes_no1):
        self.url = url  #成员变量，方便方法间使用
        self.params1 = params1
        self.headers1 = headers1
        self.method1 = method1
        self.data_type1 = data_type1
        self.check_point1 = check_point1
        self.yes_no1 = yes_no1

    def test(self):
        if self.method1 == "POST" and self.data_type1 =="Json":
           #调http请求类
            h1 = Http_request(self.url,self.params1,self.headers1)  #新建对象，传参数
            result1 = h1.post_request1()  #对象调方法，不用传参数   --字典
            print(result1)
        elif self.method1 == "POST" and self.data_type1 == "Form":
            h1 = Http_request(self.url, self.params1, self.headers1)  # 新建对象，传参数
            result1 = h1.post_request1()  # 对象调方法，不用传参数   --字典
            print(result1)
        elif self.method1 == "GET":
            h1 = Http_request(self.url, self.params1, self.headers1)  # 新建对象，传参数
            # result1 = h1.get_request1
            result1 = h1.get_request1()  # 注意2：拼写 get_request1()没有写小括号
            # 少了小括号会提示
            #  <bound method Http_request.get_request1 of <http_class2.Http_request object
            # at 0x00000000042A3518>>
            print(result1)
        else:
            print("请检查测试用例的数据")

        ret1 = result1["error_code"]
        ret2 = self.check_point1["error_code"]
        if ret1 == ret2:
            print("测试通过")
            #写入通过和响应数据到单元格
            sh1.cell(i+2, 11).value ="通过"  #这里i+2 是保证下面读取的行和写入的行是同一行
            #写入响应数据到单元格  #响应数据是字典，写入excel前需要转换成json字符串才行
            result1 = json.dumps(result1,ensure_ascii=False) #字典转换成python 字符串
            sh1.cell(i+2, 12).value = result1
            # ensure_ascii=False默认是True，不显示中文，改成False就可以显示中文
            wb1.save(path1)
        else:
            print("测试不通过")
            sh1.cell(i+2, 11).value = "不通过"  # 这里i+2 是保证下面读取的行和写入的行是同一行
            # 写入响应数据到单元格  #响应数据是字典，写入excel前需要转换成json字符串才行
            result1 = json.dumps(result1, ensure_ascii=False)  # 字典转换成python 字符串
            sh1.cell(i+2, 12).value = result1
            # ensure_ascii=False默认是True，不显示中文，改成False就可以显示中文
            wb1.save(path1)

path1 = r"D:\PycharmProjects\xiaoqiang\base_181027\a1接口层框架设计1-0326\src18\testcase2.xlsx"

#1获取工作簿
wb1 = openpyxl.load_workbook(path1)

#2获取工作表Sheet
sh1 = wb1.get_sheet_by_name("Sheet1")

#3获取最大行数
max_row1 = sh1.max_row  #没有小括号，是属性  #3

#3获取单元格
for i in range(max_row1-1):  #0,1
    url1 = sh1.cell(i+2,3).value
    # print(url1) #http://v.juhe.cn

    url2 = sh1.cell(i+2,4).value

    url = url1 +url2
    # print(url)

    method1 = sh1.cell(i+2,5).value
    # print(method1)

    data_type1 =  sh1.cell(i+2,6).value
    # print(data_type1)

    params1 = sh1.cell(i+2,7).value  #从excel读取后的字段是json字符串
    params1 = json.loads(params1)  #将json字符串转换成python字典，放入内存中
    # print(params1)
    # print(type(params1))

    check_point1 = sh1.cell(i+2,8).value#从excel读取后的字段是json字符串
    check_point1 = json.loads(check_point1) #将json字符串转换成python字典，放入内存中
    # print(check_point1)

    yes_no1 = sh1.cell(i + 2, 10).value
    # print(yes_no1)
    if yes_no1 == "否":  #如果是否运行时否，就跳出当前迭代，进行下一个迭代
        continue

    headers1 = {}

    #新建接口类对象  #注意1：这里必须放在for循环中才行，放在for循环之外之后写入最后一行
    #就是每获取一行，就调一次接口类（http请求类）
    i1 = Interface1(url,params1,headers1,method1,data_type1,check_point1,yes_no1)

    #对象调方法
    i1.test()

"""
遇到的问题总结：
1、新建接口类的对象，写在类for循环之外，会出现只写入最后一行的情况
    应该新建接口类对象和对象调方法放到for循环之中
    从而实现每获取一行excel字段，就掉一次接口类（http请求类）

2、h1.get_request1()后面少写了小括号，拼写错误
    打印result1的时候，会提示
    <bound method Http_request.get_request1 of <http_class2.Http_request object
            # at 0x00000000042A3518>>
    加上小括号之后，问题解决

方法论总结：
1、遇到报错，有足够耐心，逐行print就可以排查到原因--万能钥匙（足够耐心+逐行print）
2、不跳步骤

"""
















