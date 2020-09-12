#!/usr/bin/env python


"""
一、思路
    1、从excel读取字典，作为请求类的参数url params1 headers1，发送给服务端，获取响应数据
    2、将响应数据的错误码和检查点的错误码比对
    3、相等，将通过和响应数据写入
        不相等，将不通过和响应数据写入excel

二、结构
    1、接口类
    2、接口类中要调用Http请求类
    3、要通过判断请求方法是post还是get，请求类型是json还是form，调用请求类，等到响应数据
    4、将响应数据的错误码和检查点错误码进行比对
    5、相等，写入通过和响应数据
       不相等，写入不通过和响应数据

    6、循环遍历读取excel，获取每个字段
    7、新建接口类对象，调接口类的方法，传入实际参数

三、伪代码
    1、导包
        json
        openpyxl
        Http请求类
    2、全局变量先写死
        url，请求数据，headers1，请求方法，请求类型，检查点，是否运行
    3、接口类
        1、构造函数 传入参数，成员变量
        2、普通方法test
            1、通过判断请求方法是post还是get，请求类型是json还是form，调用http请求类（先导入文件），得到响应数据
        3、将响应数据的错误码和检查点错误码进行比对
        4、相等，写入通过和响应数据
           不相等，写入不通过和响应数据
    4、读取excel类-循环遍历case类
        循环遍历读取excel，获取每个字段
        1、指定excel文件路径
        2、获取工作簿
        3、获取工作表
        4、获取单元格
        5、新建接口类对象，调接口类的方法，传入实际参数（获取的excel字段和行号i+2）
    5、新建循环遍历case类对象，通过对象调方法
    6、新建发邮件类对象，通过对象调方法

四、注意点
    1、读取excel字段的时候，对于请求数据和检查点，读取的是json字符串，需要将json字符串转换成python的字典，放在内存中
        json.loads(params1)
        json.loads(check_point1)
    2、写入excel响应数据的时候，调用请求类，返回的是响应数据是字典，需要将字典转换成json字符串，才能写入
        json.dumps(响应数据,Ensure_ascii = False) 参数2用于显示中文
    3、写入excel后，必须wb1.save(path1),否则写入没有生效
    4、写入excel的时候，表格必须是关闭的，否则会报权限错误
    5、如何保证读取行和写入行是同一行，通过i+2

    6、从excel读取的字段，确保去掉回车换行以及两端的空格等一些无效的字符--增强健壮性和容错性
        data_type1 = sh1.cell(i+2,6).value.replace('\n','').replace('\r','').strip()

五、步骤
    1、先全局变量写死，将入参字段1
    2、然后再从excel读取一行，写入一行
    3、最后从excel读取多行，写入多行

六、扩展和优化
    错误key--1
    状态码是500 404 403等
"""

#一 导包
import json
import openpyxl
from http_class2 import Http_request1  #导入请求类文件的类
from mail_excel import Send_mail    #导入发邮件文件的类


#一、读取excel
#1指定excel的路径
# path1 = r"D:\PycharmProjects\xiaoqiang\base_181027\a1接口层框架设计1-0326\src26\testcase2.xlsx"
#绝对路径
# path1 = r".\testcase2.xlsx"  #.\表示相对路径，当前目录下，存放测试用例
# path2 = r".\testcase3.xlsx"  #相对路径，当前目录下，存放测试报告（测试用例文件填写测试结果和响应数据后）

# #2 获取工作簿
# wb1 = openpyxl.load_workbook(path2_case)
#
# #3 获取工作表
# sh1 = wb1.get_sheet_by_name("Sheet1")
#
# #3 最大行数
# max_row1 = sh1.max_row   #没有小括号，属性

#二、接口类
class Interface1:
    def test(self,url,params1,headers1,method1,data_type1,check_point1,yes_no1,i,path1_report,sh1,wb1):
        # 将读取字段，行号、测试报告路径、工作簿wb1，工作表sh1作为形式参数
        #这里的i是形参，确保读取行和写入行是同一行
        # # 2 获取工作簿
        # wb1 = openpyxl.load_workbook(path2_case)
        #
        # # 3 获取工作表
        # sh1 = wb1.get_sheet_by_name("Sheet1")

        # # 3 最大行数
        # max_row1 = sh1.max_row  # 没有小括号，属性

        h1 = Http_request1(url, params1, headers1)
        if method1 == "POST" and data_type1 == "Json":  #注意：这里的字母大小写必须和excel保持一致
            result1 = h1.post_request1()   #小括号要加上
            print(result1)
        elif method1 == "POST" and data_type1 == "Form": #注意：这里的字母大小写必须和excel保持一致
            result1 = h1.post_request1()
            print(result1)
        elif method1 == "GET":  #注意：这里的字母大小写必须和excel保持一致
            result1 = h1.get_request2()
            print(result1)
        else:
            print("请检查测试用例的数据")

        ret1 = result1["error_code"]
        ret2 = check_point1["error_code"]

        if ret1 == ret2:
            print("测试通过")
            #写入通过和响应数据（需要将响应数据从字典转换衬json字符串才能写入）
            sh1.cell(i,11).value = "通过"  #通过i+2 保证读取的行和写入的行是同一行
        else:
            print("测试不通过")
            sh1.cell(i, 11).value = "不通过"
        result1 = json.dumps(result1,ensure_ascii=False)  #字典转换成json字符串，参数2是显示中文的
        sh1.cell(i, 12).value = result1
        # wb1.save(path1)  #path1是用例文件
        wb1.save(path1_report)   #另存为到path1_report--就是一个测试报告的性质（填写了测试结果和响应数据后）


#三 获取excel字段类-遍历case类
class For_case1:
    def for_case1(self,path2_case,path1_report):  #参数1-测试用例路径 参数2-测试报告路径
        # 2 获取工作簿
        wb1 = openpyxl.load_workbook(path2_case)

        # 3 获取工作表
        sh1 = wb1.get_sheet_by_name("Sheet1")

        # 3 最大行数
        max_row1 = sh1.max_row  # 没有小括号，属性

        for i in range(max_row1-1):  #0,1

            url1 = sh1.cell(i+2,3).value.replace('\n','').replace('\r','').strip()
            #python3.3.3版本不支持（田电脑） python3.3.5版本才支持(我家里电脑)
            # url1 = sh1.cell(row=i+2,column=3).value  #python3.3.3版本只支持这个写法（田电脑）
            # print(url1)

            url2 = sh1.cell(i+2,4).value.replace('\n','').replace('\r','').strip()
            url = url1 + url2

            method1 = sh1.cell(i+2,5).value.replace('\n','').replace('\r','').strip()

            data_type1 = sh1.cell(i+2,6).value.replace('\n','').replace('\r','').strip()
            #这里是去掉从excel读取后的字段的两端的空格，以及回车换行等

            params1 = sh1.cell(i+2,7).value.replace('\n','').replace('\r','').strip()
            # params1 = json.loads(params1)  #方式1：从excel获取的是json字符串，转成python字典，放入内存
            params1 = eval(params1)  #方式2：从excel获取的是json字符串，
            # print(type(params1))


            check_point1 = sh1.cell(i+2,8).value.replace('\n','').replace('\r','').strip()
            check_point1 = json.loads(check_point1)  # 从excel获取的是json字符串，转成python字典，放入内存

            yes_no1 = sh1.cell(i+2,10).value.replace('\n','').replace('\r','').strip()
            # print(yes_no1)
            if yes_no1 == "否":
                continue  #跳出当次小循环，进行下一次小循环

            headers1 = {}  #容易忘记 注意

            # 新建接口类对象，普通方法传入参数，  （调接口类方法，必须在for内，每读取一行，发一次请求）
            i1 = Interface1()
            i1.test(url, params1, headers1, method1, data_type1, check_point1, yes_no1, i+2,path1_report,sh1,wb1)
            #将读取字段，行号、测试报告路径、工作簿wb1，工作表sh1作为实际参数，传递给接口类

# #五、新建遍历case类对象
# f1 = For_case1()
# #对象调方法
# f1.for_case1()
#
# #六、新建发邮件类对象
# s1 = Send_mail()
# #对象调方法
# s1.send_mail("接口测试报告190402")
# print("邮件发送成功")

"""
总结：
1、当类中的普通方法只有1个的时候，是不需要构造方法的，至少2个以上普通方法，才需要构造方法
2、循环case类中，没读取一行，就调用一次接口类的方法，将i+2的行数作为实参，传入到接口类的test方法

1、遍历case类，想把参数（从excel读取的字段以及行号）传递给接口类的办法是下面2个：
    方法1、直接通过接口类的对象调接口类的方法，这些参数作为实际参数传递
       同时在接口类的普通方法，写上这些参数的形式参数即可---推荐
       这么做的好处，每遍历一行，就调用一次接口类的方法，非常适合for循环--重点
    方法2、如果不是for循环，而是只有一行excel字段以及行号
        还可以考虑，把这些参数作为元组，return，通过多个变量接收这个元组
        1、先新建遍历case类的对象，对象调方法，return元组
        2、用多个变量接收return的元组
        3、这些变量，放在接口类中
        4、新建接口类对象，调接口类的方法
        方法2不适合for循环的情况
    小结：01for循环中，每读取一行就掉一次其他类的方法--适合for循环
          02如果不是for循环，不要求每读取一行就掉一次其他类的方法，--就可以用return元组
          新建类1的对象，对象调方法，通过多个变量接收return的元组即可

2、知识点总结
    从excel读取的字段--请求数据、检查点是json字符串类型，需要转换成python字典
    下面两种方法都可以实现
     params1 = json.loads(params1)  #方式1：从excel获取的是json字符串，转成python字典，放入内存
     params1 = eval(params1)

3、目录结构：
    1、公共类库目录下放1个文件
        发送邮件文件--mail_excel.py
    2、src目录下放3个文件
        1、主入口文件-mail1.py
        2、请求类文件-http_class2
        3、接口类和循环遍历case类文件-a1_interface2.py
            --一个文件存放了2个类

"""





























