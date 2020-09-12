#!/usr/bin/env python
#-*- coding:utf-8 -*-

import json
import openpyxl
from Http_request1 import Http_request2

#1全局变量
# url = "http://v.juhe.cn/laohuangli/d"
# params1 = {"key": "e711bc6362b3179f5a28de7fd3ee4ace", "date": "2016-5-14"}
# headers1 = {}
# method1 = "POST"
# data_type1 = "Json"
# check_point1 = {"error_code":0}
# params1 = {"key": "e711bc6362b3179f5a28de7fd3ee4ace", "date": "2016-5-14"}
# yes_no1 = "是"

# 3读取excel
path1 = r"D:\xq_py\06接口自动化框架\testcase.xlsx"
wb1 = openpyxl.load_workbook(path1)
# print(wb1)
sh1 = wb1.get_sheet_by_name("Sheet1")
max_row1 = sh1.max_row
# print(max_row1)

#2接口类
class Interface2:
    def __init__(self,url,headers1,params1,method1,data_type1,yes_no1,check_point1):
        self.url = url
        self.headers1 = headers1
        self.params1 = params1
        self.method1 = method1
        self.data_type1 = data_type1
        self.yes_no1 = yes_no1
        self.check_point1 = check_point1

    def test(self):
        if self.method1 == "POST" and self.data_type1 == "Json":
            h1 = Http_request2(self.url,self.params1,self.headers1)
            result1 = h1.post_request1()
            print(result1)

        ret1 = result1["error_code"]
        ret2 = self.check_point1["error_code"]

        result1 = json.dumps(result1,ensure_ascii=False)
        for i in ret[0]:
            if ret1 == ret2:
                print("测试通过")
                sh1.cell(row=i,column=11).value = "通过"
                sh1.cell(row=i, column=12).value = result1
                wb1.save(path1)
            else:
                print("测试不通过")

class For_case:
    def for_case(self):
        #3读取excel
        # path1 = r"D:\xq_py\06接口自动化框架\testcase.xlsx"
        # wb1 = openpyxl.load_workbook(path1)
        # # print(wb1)
        # sh1 = wb1.get_sheet_by_name("Sheet1")
        # # print(sh1)
        li1 = []
        li2 = []
        for i in range(max_row1-1):
            url1 = sh1.cell(row=i+2,column=3).value  #py3.3只支持这个
            # url1 = sh1.cell(3,3).value   #py3.7可以支持这个
            # print(url1)

            url2 = sh1.cell(row=i+2,column=4).value
            url = url1 +url2

            method1 = sh1.cell(row=i+2,column=5).value

            data_type1 = sh1.cell(row=i+2,column=6).value

            params1 = sh1.cell(row=i+2,column=7).value
            params1 = json.loads(params1)

            check_point1 = sh1.cell(row=i+2,column=8).value
            check_point1 = json.loads(check_point1)

            yes_no1 = sh1.cell(row=i+2,column=10).value
            if yes_no1 =="否":
                continue

            headers1 ={}

            li1.append(i+2)
            li2.append((url,headers1,params1,method1,data_type1,yes_no1,check_point1))

            # print(i + 2)
            # return(url,headers1,params1,method1,data_type1,yes_no1,check_point1,i+2)
            # return(i+2)
        # print(li1)
        # print(li2)
        return (li1,li2)

#4新建循环case类对象
f1 = For_case()
ret = f1.for_case()
# print(ret)
# print(ret[0])
# print(ret[1])
# print(type(ret[1]))

#5新建接口类对象
# i1 = Interface2(url,headers1,params1,method1,data_type1,yes_no1,check_point1)
for i in ret[1]:
    i1 = Interface2(*i)
    i1.test()














